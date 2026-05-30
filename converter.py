"""
converter.py — DR CSV to Excel Converter
==========================================
Reads SAS-exported CSV files and produces a formatted Excel file
matching your DR template exactly. Eliminates copy-paste entirely.

SETUP (one time only):
  1. Install requirements:
       pip install openpyxl pandas
  2. Create the templates folder (already exists if you extracted the ZIP):
       DR_Generator/templates/
  3. Copy your Excel templates into that folder:
       DR_Template_CrashOnly.xlsx
       DR_Template_CVO.xlsx
       DR_Template_CVO_GRP.xlsx
       DR_Template_VehOnly.xlsx
       DR_Template_CrashOnly_GRPData.xlsx
       DR_Template_C_V.xlsx

USAGE (from VS Code terminal — Ctrl+` to open terminal):
  python converter.py "R:\\DR-2847-Smith" 2847

  Template type is auto-detected from which CSV files are in the folder.
  Output is saved as DR2847_Data.xlsx in that same folder.

MANUAL OVERRIDE (if auto-detect picks the wrong template):
  python converter.py "R:\\DR-2847-Smith" 2847 cvo
  python converter.py "R:\\DR-2847-Smith" 2847 crash
  python converter.py "R:\\DR-2847-Smith" 2847 crash_grp
  python converter.py "R:\\DR-2847-Smith" 2847 cvo_grp
  python converter.py "R:\\DR-2847-Smith" 2847 veh
  python converter.py "R:\\DR-2847-Smith" 2847 cv
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path

# ============================================================
# CONFIGURATION
# ============================================================

SCRIPT_DIR = Path(__file__).parent

# Where your Excel templates live.
# Change this if you keep them on a shared drive instead.
TEMPLATES_DIR = SCRIPT_DIR / 'templates'

# Maps template type → Excel template filename.
# Update these names if your template files are named differently.
TEMPLATE_FILES = {
    'crash':     'DR_Template_CrashOnly.xlsx',
    'crash_grp': 'DR_Template_CrashOnly_GRPData.xlsx',
    'cvo':       'DR_Template_CVO.xlsx',
    'cvo_grp':   'DR_Template_CVO_GRP.xlsx',
    'veh':       'DR_Template_VehOnly.xlsx',
    'cv':        'DR_Template_C_V.xlsx',
}

# Maps template type → which CSV goes into which sheet.
# Key = csv type ('crash'/'veh'/'occ'), Value = sheet name in Excel template.
SHEET_CONFIG = {
    'crash':     {'crash': 'DR#'},
    'crash_grp': {'crash': 'DR#'},
    'cvo':       {'crash': 'DR_CRASH', 'veh': 'DR_VEH', 'occ': 'DR_OCC'},
    'cvo_grp':   {'crash': 'DR_CRASH', 'veh': 'DR_VEH', 'occ': 'DR_OCC'},
    'veh':       {'veh':   'DR_VEH'},
    'cv':        {'crash': 'DR_CRASH', 'veh': 'DR_VEH'},
}


# ============================================================
# STEP 1 — FIND CSV FILES
# ============================================================

def find_csvs(folder, dr_num):
    """
    Scans the folder for SAS-exported CSV files for this DR number.
    Returns a dict like: {'crash': Path(...), 'veh': Path(...), 'occ': Path(...)}
    """
    folder = Path(folder)
    found = {}
    dr_tag = str(dr_num).upper()

    for f in sorted(folder.iterdir()):
        if not f.suffix.upper() == '.CSV':
            continue
        name = f.stem.upper()

        # Must contain the DR number
        if 'DR' + dr_tag not in name.replace('-', '').replace('_', '').replace(' ', ''):
            # Also check just the number itself in case named differently
            if dr_tag not in name:
                continue

        if 'CRASH' in name and 'crash' not in found:
            found['crash'] = f
        elif 'VEH' in name and 'veh' not in found:
            found['veh'] = f
        elif 'OCC' in name and 'occ' not in found:
            found['occ'] = f

    return found


# ============================================================
# STEP 2 — DETECT TEMPLATE TYPE
# ============================================================

def detect_type(found_csvs):
    """
    Auto-detects which template to use based on which CSVs exist
    and whether the crash CSV has a GRP/REQUESTED INTERSECTION column.
    """
    has_crash = 'crash' in found_csvs
    has_veh   = 'veh'   in found_csvs
    has_occ   = 'occ'   in found_csvs

    # Check for GRP data by peeking at crash CSV header
    is_grp = False
    if has_crash:
        try:
            header = pd.read_csv(found_csvs['crash'], nrows=0)
            first_col = header.columns[0].upper() if len(header.columns) > 0 else ''
            is_grp = 'REQUESTED' in first_col or first_col.startswith('GRP')
        except Exception:
            pass

    if has_crash and has_veh and has_occ:
        return 'cvo_grp' if is_grp else 'cvo'
    elif has_crash and has_veh:
        return 'cv'
    elif has_crash:
        return 'crash_grp' if is_grp else 'crash'
    elif has_veh:
        return 'veh'
    else:
        raise ValueError(
            f'Could not detect template type. '
            f'Found these CSV types: {list(found_csvs.keys())}'
        )


# ============================================================
# STEP 3 — MAP CSV COLUMNS TO TEMPLATE COLUMNS
# ============================================================

def build_column_map(template_headers, csv_headers):
    """
    Maps CSV columns to template column positions.

    Strategy:
      1. Try exact name match (case-insensitive, trimmed)
      2. If no match, fall back to positional (col 1 = col 1, col 2 = col 2, ...)

    Returns a list of (template_col_index, csv_col_index) pairs — 0-based.
    """
    t_headers = [str(h).strip().upper() if h else '' for h in template_headers]
    c_headers = [str(h).strip().upper() if h else '' for h in csv_headers]

    mapping = []
    unmatched_template = []
    unmatched_csv = set(range(len(c_headers)))

    # Try name matching first
    for t_idx, t_name in enumerate(t_headers):
        if not t_name:
            continue
        matched = False
        for c_idx, c_name in enumerate(c_headers):
            if t_name == c_name:
                mapping.append((t_idx, c_idx))
                unmatched_csv.discard(c_idx)
                matched = True
                break
        if not matched:
            unmatched_template.append(t_idx)

    # If fewer than half matched by name, fall back to positional
    if len(mapping) < len(t_headers) * 0.5:
        mapping = []
        for i in range(min(len(t_headers), len(c_headers))):
            mapping.append((i, i))
        print('    Note: using positional column mapping (header names did not match).')
    elif unmatched_template:
        print(f'    Note: {len(unmatched_template)} template columns had no matching CSV column '
              f'(will be left blank).')

    return mapping


# ============================================================
# STEP 4 — WRITE DATA INTO TEMPLATE SHEET
# ============================================================

def write_to_sheet(ws, df, verbose=True):
    """
    Writes dataframe rows into a worksheet starting at row 2.
    Row 1 (template headers) is preserved unchanged.
    Columns are mapped by name matching with positional fallback.
    """
    # Get template headers from row 1
    template_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Build column map
    col_map = build_column_map(template_headers, list(df.columns))

    # Write data rows
    for row_num, row_data in enumerate(df.itertuples(index=False), start=2):
        row_list = list(row_data)
        for t_col, c_col in col_map:
            if c_col >= len(row_list):
                continue
            val = row_list[c_col]
            # Write None for NaN/empty values
            if pd.isna(val) if isinstance(val, float) else val is None:
                ws.cell(row_num, t_col + 1).value = None
            else:
                ws.cell(row_num, t_col + 1).value = str(val) if val != '' else None

    if verbose:
        print(f'    {len(df)} rows written across {len(col_map)} columns.')


# ============================================================
# STEP 5 — UPDATE DATA NOTES SHEET
# ============================================================

def update_data_notes(wb, dr_num):
    """
    Updates the DR# placeholder in the Data_Notes sheet
    with the actual DR number.
    """
    if 'Data_Notes' not in wb.sheetnames:
        return
    ws = wb['Data_Notes']
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'DR#' in cell.value:
                    cell.value = cell.value.replace('DR#', f'DR{dr_num}')


# ============================================================
# MAIN CONVERT FUNCTION
# ============================================================

def convert(folder, dr_num, template_type=None):
    """
    Main function — orchestrates the full conversion.

    folder:        path to the DR folder containing CSV files
    dr_num:        DR number as a string (e.g. '2847')
    template_type: optional override; auto-detected if None
    """
    folder = Path(folder)

    print(f'\n{"="*55}')
    print(f'  DR Converter')
    print(f'  DR Number : {dr_num}')
    print(f'  Folder    : {folder}')
    print(f'{"="*55}')

    if not folder.exists():
        raise FileNotFoundError(f'Folder not found: {folder}')

    # Find CSV files
    csvs = find_csvs(folder, dr_num)
    if not csvs:
        raise FileNotFoundError(
            f'\nNo CSV files found for DR{dr_num} in:\n  {folder}\n\n'
            f'Expected filenames like:\n'
            f'  DR{dr_num}_CrashData.csv\n'
            f'  DR{dr_num}_VehData.csv\n'
            f'  DR{dr_num}_OccData.csv'
        )
    print(f'\nFound CSVs:')
    for k, v in csvs.items():
        print(f'  {k:8s} → {v.name}')

    # Detect or use provided template type
    if template_type:
        ttype = template_type.lower().replace('-', '_')
        if ttype not in TEMPLATE_FILES:
            raise ValueError(
                f'Unknown template type: "{template_type}"\n'
                f'Valid options: {", ".join(TEMPLATE_FILES.keys())}'
            )
    else:
        ttype = detect_type(csvs)

    print(f'\nTemplate type : {ttype}')

    # Locate template file
    template_path = TEMPLATES_DIR / TEMPLATE_FILES[ttype]
    if not template_path.exists():
        raise FileNotFoundError(
            f'\nTemplate file not found:\n  {template_path}\n\n'
            f'Please copy your Excel templates into:\n  {TEMPLATES_DIR}\n\n'
            f'Required file: {TEMPLATE_FILES[ttype]}'
        )
    print(f'Template file : {template_path.name}')

    # Load workbook
    wb = load_workbook(template_path)
    sheet_config = SHEET_CONFIG[ttype]

    # Process each sheet
    print()
    for csv_key, sheet_name in sheet_config.items():
        if csv_key not in csvs:
            print(f'  SKIP  "{sheet_name}" — no {csv_key} CSV found')
            continue

        if sheet_name not in wb.sheetnames:
            print(f'  ERROR "{sheet_name}" sheet not found in template')
            continue

        print(f'  Writing "{sheet_name}" from {csvs[csv_key].name} ...')
        df = pd.read_csv(
            csvs[csv_key],
            dtype=str,           # keep everything as text — SAS already formatted values
            encoding='utf-8-sig' # handles BOM from SAS CSV exports
        )
        df = df.fillna('')

        ws = wb[sheet_name]
        write_to_sheet(ws, df)

    # Update Data_Notes with actual DR number
    update_data_notes(wb, dr_num)

    # Save output
    out_path = folder / f'DR{dr_num}_Data.xlsx'
    wb.save(out_path)

    print(f'\n{"="*55}')
    print(f'  Output saved:')
    print(f'  {out_path}')
    print(f'{"="*55}\n')

    return out_path


# ============================================================
# COMMAND LINE ENTRY POINT
# ============================================================

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    folder_arg  = sys.argv[1]
    dr_num_arg  = sys.argv[2]
    type_arg    = sys.argv[3] if len(sys.argv) > 3 else None

    try:
        convert(folder_arg, dr_num_arg, type_arg)
    except FileNotFoundError as e:
        print(f'\n  ERROR: {e}\n')
        sys.exit(1)
    except Exception as e:
        print(f'\n  ERROR: {e}\n')
        import traceback
        traceback.print_exc()
        sys.exit(1)
